
import openpyxl
from openpyxl.styles import Font
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from scipy import stats
from matplotlib.colors import LinearSegmentedColormap
import os


def load_workbook(file_path):
    return openpyxl.load_workbook(file_path)


def calculate_score(value, thresholds):
    if value is None:
        return 1  # Default value for missing data

    # Extract only the numeric thresholds
    numeric_thresholds = [t for t in thresholds if isinstance(t, (int, float)) and t is not None]
    
    if not numeric_thresholds:
        return 1  # Default value if no valid thresholds are found

    # Determine if higher or lower values are better
    higher_is_better = numeric_thresholds[0] < numeric_thresholds[-1]

    if higher_is_better:
        sorted_thresholds = sorted(numeric_thresholds)
        for i, threshold in enumerate(sorted_thresholds):
            if value <= threshold:
                return i + 1
        return 5  # If value is GREATER than all thresholds
    else:
        sorted_thresholds = sorted(numeric_thresholds, reverse=True)
        for i, threshold in enumerate(sorted_thresholds):
            if value >= threshold:
                return i + 1
        return 5  # If value is lower than all thresholds

  
def process_sheet(sheet):
    total_score = 0
    total_weight = 0
    weighted_scores = {}
    
 
    index = 1
    for row in sheet.iter_rows(min_row=2, max_col=17):
        index += 1
        if not row[4].value:  # Stop if we reach an empty row in column E
            break
        
        weight = row[3].value  # Column D
        param = row[4].value   # Column E
        value = row[5].value   # Column F
        thresholds = [cell.value for cell in row[6:11]]  # Columns G to K
        
        if weight is not None:
            score = calculate_score(value, thresholds)
            weighted_score = score * weight
            weighted_scores[param] = weighted_score

            total_score += weighted_score
            total_weight += weight  # Write individual score to Output column L



            #print(f"Parameter: {param}, Value: {value}, Score: {score}")  # Debug print


    # Calculate final score out of 100
    #final score as sum of weighted scores
    if(round(total_weight, 3) != 1.00):
                print(f"The total weight was {round(total_weight, 3)}, it should be 1.00. Please adjust accordingly")
            #error if not 100
    #final_score = (total_score / total_weight) * 20 if total_weight > 0 else 0
    final_score = total_score * 20 if total_weight > 0 else 0

    # Write final score at the bottom of the Output column
    last_row = index + 1
    final_score_cell = sheet.cell(row=last_row, column=12)  # Column L?
    final_score_cell.value = final_score
    final_score_cell.font = Font(bold=True)
    sheet.cell(row=last_row, column=5, value="Final Score:").font = Font(bold=True)
    print(final_score)

 
    return final_score, weighted_scores





def plot_normal_distribution(scores, plots):
    df = pd.DataFrame(scores, columns=['Score'])
    
    # Calculate mean and standard deviation
    mean = df['Score'].mean()
    std = df['Score'].std()

    # Create a range of x values
    x = np.linspace(mean - 3*std, mean + 3*std, 100)

    # Calculate the normal distribution
    y = stats.norm.pdf(x, mean, std)

    # Plot the histogram and the PDF
    plt.figure(figsize=(10, 6))
    #df['Score'].hist(density=True, bins=20, alpha=1.0, color='olivedrab')
        # Calculate y-coordinates for scatter points using the PDF
    


    plt.plot(x, y, color='green', marker=',', linestyle='dotted', linewidth=2, markersize=10)

    y_scatter = stats.norm.pdf(df['Score'], mean, std)
    
    # Scatter plot
    plt.scatter(df['Score'], y_scatter, alpha=1.0, color='blue')
    #plt.plot(x, y, 'r-', lw=2)



    
    plt.title('Normal Distribution of Scores')
    plt.xlabel('Score')
    plt.ylabel('Density')
    
    # Add mean and std dev to the plot
    plt.axvline(mean, color='deepskyblue', linestyle='dashed', linewidth=2)
    plt.text(mean*1.1, plt.ylim()[1]*0.9, f'Mean: {mean:.2f}', color='blue')
    plt.text(mean*1.1, plt.ylim()[1]*0.8, f'SD: {std:.2f}', color='blue')

    print("1")
    plt.tight_layout()
    print("2")
    plot_path = os.path.abspath('myplot.png')
    print("3")
    plt.savefig(plot_path, dpi=75, bbox_inches='tight')
    print("4")
    
    
    img = openpyxl.drawing.image.Image(plot_path)
    print("5")
    img.anchor = 'A1'
    print("6")
    plots.add_image(img)
    print("7")
    #plt.show()
    print("8")

def plot_scatter(scores, plots):
    df = pd.DataFrame(scores, columns=['Score'])
    
    plt.figure(figsize=(10, 6))
    
    # Normalize scores to use as sizes
    min_score = df['Score'].min()
    max_score = df['Score'].max()
    normalized_scores = (df['Score'] - min_score) / (max_score - min_score)
    
    # Set size range: smallest point 20, largest 200
    sizes = normalized_scores * 180 + 20
    
    # Generate random colors for each point
    colors = [mcolors.rgb2hex(np.random.rand(3)) for _ in range(len(df))]
    
    plt.scatter(range(len(df)), df['Score'], alpha=1.0, c=colors, s=sizes)
    
    plt.title('Scatter Plot of Scores')
    plt.xlabel('Index')
    plt.ylabel('Score')
    
    mean = df['Score'].mean()
    std = df['Score'].std()
    
    plt.axhline(mean, color='blue', linestyle='dashed', linewidth=2)
    
    plt.text(0.95, 0.95, f'Mean: {mean:.2f}', transform=plt.gca().transAxes, 
             verticalalignment='top', horizontalalignment='right', color='blue')
    plt.text(0.95, 0.90, f'SD: {std:.2f}', transform=plt.gca().transAxes, 
             verticalalignment='top', horizontalalignment='right', color='blue')
    
    plt.tight_layout()
    plot_path = os.path.abspath('scatter_plot.png')
    plt.savefig(plot_path, dpi=75, bbox_inches='tight')
    
    img = openpyxl.drawing.image.Image(plot_path)
    img.anchor = 'M1'  # This will place the second graph next to the first one
    plots.add_image(img)
    plt.close()

'''
def plot_pie_chart(sheet_name, weighted_scores, sheet):
    plt.figure(figsize=(10, 6))
    labels = list(weighted_scores.keys())
    sizes = list(weighted_scores.values())
    
    plt.pie(sizes, labels=labels, autopct='%1.1f%%', textprops={'size': 'smaller'}, startangle=90)
    plt.axis('equal')
    plt.title(f'Weighted Scores Distribution - {sheet_name}')
    
    plot_path = os.path.abspath(f'pie_chart_{sheet_name}.png')
    plt.savefig(plot_path, dpi=75, bbox_inches='tight')
    
    img = openpyxl.drawing.image.Image(plot_path)
    img.anchor = 'N1'  # This will place the pie chart in column N, row 1
    sheet.add_image(img)
    plt.close()
    #os.remove(plot_path)  # Remove the temporary image file
'''

def plot_pie_chart(sheet_name, weighted_scores, sheet):
    plt.figure(figsize=(12, 8))
    labels = list(weighted_scores.keys())
    sizes = list(weighted_scores.values())
    
    # Calculate total score and points lost
    total_score = sum(sizes)
    points_lost = 5.0 - total_score

    # Add points lost to the data
    sizes.append(points_lost)
    labels.append("Points Lost")

    # Sort data by size
    sorted_data = sorted(zip(sizes, labels), reverse=True)
    sizes, labels = zip(*sorted_data)
    
    # Group small slices
    threshold = 0.07  # Adjust as needed
    other_sum = sum(size for size in sizes if size < threshold)

    sizes = [size for size in sizes if size >= threshold] + [other_sum]
    labels = [label for size, label in zip(sorted_data, labels) if size[0] >= threshold] + ["Other"] 
    
    # Create a custom color map with more unique colors
    n_colors = len(sizes)
    colors = list(plt.cm.tab20(np.arange(n_colors) / n_colors))

    # Find the index of "Points Lost" and set its color to black
    points_lost_index = labels.index("Points Lost") if "Points Lost" in labels else -1
    if points_lost_index != -1:
        colors[points_lost_index] = 'black'
    
    # Create pie chart
    wedges, texts, autotexts = plt.pie(sizes, labels=None, autopct='%1.1f%%', 
                                       colors=colors, startangle=90, 
                                       pctdistance=0.85, labeldistance=1.05)
    
    # Add a legend
    plt.legend(wedges, labels, title="Categories", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    plt.title(f'Weighted Scores Distribution - {sheet_name}')
    plt.axis('equal')
    
    # Adjust layout to prevent clipping of labels
    plt.tight_layout()
    
    plot_path = os.path.abspath(f'pie_chart_{sheet_name}.png')
    plt.savefig(plot_path, dpi=90, bbox_inches='tight')
    
    img = openpyxl.drawing.image.Image(plot_path)
    img.anchor = 'N1'  # This will place the pie chart in column N, row 1
    sheet.add_image(img)
    plt.close()
    #os.remove(plot_path)  # Remove the temporary image file

def clear_plots_sheet(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None
    for img in sheet._images:
        sheet._images.remove(img)

def clear_plots(sheet):
    for img in sheet._images:
        sheet._images.remove(img)


def main():

    file_path = 'python_testsite.xlsx'  # Replace with your actual file path
    workbook = load_workbook(file_path)
    final_scores = []
    all_scores = []
    plots = workbook['Plots']
    clear_plots_sheet(plots)

 
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if sheet_name == 'Plots':
            continue
        clear_plots(sheet)
        final_score, weighted_scores = process_sheet(sheet)
        final_scores.append(final_score)
        # Plot pie chart for each sheet
        plot_pie_chart(sheet_name, weighted_scores, sheet)

        print(f"Final Score for sheet '{sheet_name}': {final_score:.2f}/100")
 
    all_scores.extend(final_scores)



    plot_normal_distribution(all_scores, plots)
    plot_scatter(all_scores, plots)


        
    workbook.save(file_path)

if __name__ == "__main__":
    main()

